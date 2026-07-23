"""Build template/wif_workbook.xlsx (README + RULES) for the WIF kernel.

Run from the repo root:  python tools/make_wif_workbook.py
Regenerates the committed template deterministically.
"""
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "template", "wif_workbook.xlsx")
OUT = os.path.abspath(OUT)
os.makedirs(os.path.dirname(OUT), exist_ok=True)

HDR_FILL = PatternFill("solid", fgColor="205081")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
NOTE_FONT = Font(italic=True, color="666666", name="Calibri", size=10)
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="205081")
H2_FONT = Font(bold=True, size=11, name="Calibri", color="205081")
BODY_FONT = Font(name="Calibri", size=11)
COMMENT_FONT = Font(italic=True, color="808080", name="Consolas", size=10)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------------- README ----------------
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = "205081"
ws.column_dimensions["A"].width = 116
rows = [
    ("T", "WIF - hook-based what-if rules"),
    ("B", ""),
    ("B", "One sheet (RULES) defines everything. In SAS:"),
    ("B", '    %include "<server-path>/wif.sas";'),
    ("B", '    %wif_init(scenario=RENEWAL, rules=<this workbook: SERVER path>);'),
    ("B", "    ... run your program (its %wif() hooks fire) ...  %wif_off;"),
    ("B", ""),
    ("H", "The RULES columns"),
    ("B", "  scenario   which scenario owns the rule. BLANK = global (applies to every scenario)."),
    ("B", "  hook       WHERE the rule fires: the table name hooked by %wif(table), a custom name"),
    ("B", "             matched by %wif(table, at=NAME), or INPUT (applied to a raw input table"),
    ("B", "             at %wif_init time, via staged copies - the base folder is never touched)."),
    ("B", "  seq        order within the hook (10, 20, 30 ... leave blank for sheet order)."),
    ("B", "  active     Y/N. N rows are ignored."),
    ("B", "  verb       SET / JOIN / FILTER / APPEND / REPLACE / CODE / ASSERT / SAVE /"),
    ("B", "             SORT / DEDUPE / LET (reference below)."),
    ("B", "  target     usually BLANK = the hooked table itself. INPUT rules need libref.table."),
    ("B", "  where_clause  row condition. SET/JOIN: IF syntax (in, =:, missing(), date literals);"),
    ("B", "             FILTER/APPEND: full WHERE syntax (LIKE, BETWEEN, IS MISSING ...)."),
    ("B", "  keys       JOIN only: space-separated key columns (same names both sides)."),
    ("B", "  source     JOIN/APPEND/REPLACE: the lookup/donor table (WORK.X or libref.X)."),
    ("B", "  columns    JOIN/APPEND/REPLACE column mapping: srccol or srccol=targetcol, spaces"),
    ("B", "             between. JOIN with columns blank = update same-named columns."),
    ("B", "  assign     SET/JOIN: SAS assignment statements. CODE: the code. LET: the value."),
    ("B", "  options    ONCE  ITERS=1|2+|ALL  NEWCOLS  NOWARN0  DROP  KEEPEXTRA  ALLOWLIB"),
    ("B", "             LAST (DEDUPE)  NOMATCH=KEEP|FAIL|DELETE (JOIN)"),
    ("B", "  notes      free text for humans."),
    ("B", ""),
    ("H", "Verb reference"),
    ("B", "  SET      change columns in place:  assign 'policy_age = policy_age + 1;'"),
    ("B", "  JOIN     the workhorse: hash left-join source by keys; matched rows can pull columns"),
    ("B", "           (columns cell) AND compute (assign cell: 'rate = rate * adj_factor;')."),
    ("B", "           Unmatched rows are untouched; duplicate source keys are a hard error."),
    ("B", "  FILTER   keep rows matching where_clause (options DROP inverts)."),
    ("B", "  APPEND   add the source's rows (where_clause filters the SOURCE)."),
    ("B", "  REPLACE  swap the whole table for the source (trimmed to the target's columns"),
    ("B", "           unless KEEPEXTRA)."),
    ("B", "  ASSERT   QA gate, never modifies: where_clause = the condition EVERY row must"),
    ("B", "           satisfy (full WHERE). Violations -> FAILED + sample in work.wif_viol."),
    ("B", "  SAVE     snapshot the hooked table; the DESTINATION goes in the source column"),
    ("B", "           (parameters fine: RESULTS.SNAP_&WIF_SCENARIO.)."),
    ("B", "  SORT     keys with DESCENDING allowed.   DEDUPE  keep FIRST (or LAST) per keys."),
    ("B", "  CODE     escape hatch: the assign cell is inlined verbatim as generated code."),
    ("B", "           You may reference the live built-ins WIF_TABLE / WIF_HOOK there."),
    ("B", "  LET      define a parameter: target = its name, assign = its value. Rows with a"),
    ("B", "           scenario override global rows of the same name. Reference as &NAME."),
    ("B", ""),
    ("H", "Cell rules (the lint enforces all of these before anything runs)"),
    ("B", "  * Text literals in single quotes: region = 'EAST', dates '01JAN2027'd."),
    ("B", "  * Ampersand / percent inside single quotes is fine ('R&D', '5%')."),
    ("B", "  * &NAME outside quotes must be a LET parameter, params= value, or a built-in"),
    ("B", "    (&WIF_ITER, &WIF_SCENARIO; &WIF_TABLE / &WIF_HOOK inside CODE cells)."),
    ("B", "  * No /* comments in cells - use notes. Keep expression columns Text-formatted."),
    ("B", "  * A row whose scenario or hook cell starts with # is a comment row."),
    ("B", ""),
    ("H", "Remember"),
    ("B", "  * Hooks are inert until %wif_init runs - safe to leave in the program permanently."),
    ("B", "  * INPUT rules stage modified copies under WORK; your base folder stays pristine."),
    ("B", "  * work.wif_log records every rule application (rows before/after/affected)."),
    ("B", "  * Save this workbook where the SAS SERVER can read it (EG users: server path!)."),
]
r = 1
for kind, text in rows:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "T":
        c.font = TITLE_FONT
    elif kind == "H":
        c.font = H2_FONT
    else:
        c.font = BODY_FONT
    c.alignment = Alignment(wrap_text=False, vertical="top")
    r += 1

# ---------------- RULES ----------------
ws = wb.create_sheet("RULES")
ws.sheet_properties.tabColor = "2E7D32"
COLS = [
    ("scenario", 14), ("hook", 16), ("seq", 6), ("active", 8), ("verb", 10),
    ("target", 16), ("where_clause", 34), ("keys", 20), ("source", 18),
    ("columns", 26), ("assign", 44), ("options", 16), ("notes", 34),
]
for j, (name, width) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(j)].width = width

EXAMPLES = [
    ("# --- RENEWAL: the policy one term later (used in the CLV loop driver) ---",
     "", None, "", "", "", "", "", "", "", "", "", ""),
    ("RENEWAL", "POLICIES", 10, "Y", "SET", "", "", "", "", "",
     "policy_age = policy_age + 1;", "", "everyone is one year older at renewal"),
    ("RENEWAL", "GUIDANCE_100", 10, "Y", "JOIN", "", "", "POL_ID", "WORK.CARRY",
     "SCHED_MOD=EXPIRING_MOD", "", "ITERS=2+",
     "prior term's schedule mod becomes the expiring mod (driver builds WORK.CARRY)"),
    ("RENEWAL", "GUIDANCE_100", 20, "Y", "SET", "", "", "", "", "",
     "rate_change = rate_change * &RATE_BUMP;", "", "rate change drift assumption"),
    ("RENEWAL", "", None, "Y", "LET", "RATE_BUMP", "", "", "", "", "1.05", "",
     "parameter used above; override per scenario or via params="),
    ("# --- MKTADJ: bulk market adjustment by segment, plus an INPUT tweak ---",
     "", None, "", "", "", "", "", "", "", "", "", ""),
    ("MKTADJ", "RATED", 10, "Y", "JOIN", "", "", "NAICS LOB STATE", "WORK.MKT_ADJ",
     "ADJ_FACTOR", "rate = rate * adj_factor;", "",
     "left-join the adjustment table you built; columns listed there may be new"),
    ("MKTADJ", "INPUT", 10, "Y", "SET", "RAW.RATES", "rate_year = 2027", "", "", "",
     "rate_change = rate_change * 1.10;", "",
     "modifies a STAGED COPY of the raw input at wif_init; base folder untouched"),
]
for i, row in enumerate(EXAMPLES, 2):
    is_comment = str(row[0]).startswith("#")
    for j, val in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = COMMENT_FONT if is_comment else BODY_FONT
        c.border = BORDER
        if COLS[j - 1][0] != "seq":
            c.number_format = "@"

# Text format for the whole entry area (rows 2..500) except seq
for j, (name, _) in enumerate(COLS, 1):
    if name == "seq":
        continue
    for i in range(2, 501):
        ws.cell(row=i, column=j).number_format = "@"

dv_verb = DataValidation(
    type="list",
    formula1='"SET,JOIN,FILTER,APPEND,REPLACE,CODE,ASSERT,SAVE,SORT,DEDUPE,LET"',
    allow_blank=True, showErrorMessage=False)
dv_act = DataValidation(
    type="list", formula1='"Y,N"', allow_blank=True, showErrorMessage=False)
ws.add_data_validation(dv_verb)
ws.add_data_validation(dv_act)
dv_verb.add("E2:E500")
dv_act.add("D2:D500")
ws.freeze_panes = "A2"

wb.save(OUT)
print("wrote", OUT)

# ---------------- round-trip assertions ----------------
chk = load_workbook(OUT)
assert chk.sheetnames == ["README", "RULES"], chk.sheetnames
rs = chk["RULES"]
hdr = [rs.cell(row=1, column=j).value for j in range(1, len(COLS) + 1)]
assert hdr == [n for n, _ in COLS], hdr
assert rs.cell(row=2, column=1).value.startswith("#"), "comment row lost"
assert rs.cell(row=3, column=11).value == "policy_age = policy_age + 1;"
assert rs.cell(row=3, column=11).number_format == "@", "assign not Text-formatted"
assert rs.cell(row=6, column=5).value == "LET"
assert len(rs.data_validations.dataValidation) == 2, "dropdowns missing"
assert rs.freeze_panes == "A2"
print("round-trip assertions passed")
